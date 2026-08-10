"""Read-only demographics table parsing and explicit subject alignment."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from neuroagent.application.errors import InputValidationError
from neuroagent.infrastructure.filesystem.dataset_inspector import sha256_file


def _validated_header(values: list[Any], *, source: str) -> list[str]:
    header = [str(value).strip() if value is not None else "" for value in values]
    if not header or not all(header):
        raise InputValidationError(
            "demographics_header_invalid",
            f"{source} 表头不能为空。",
        )
    if len(set(header)) != len(header):
        raise InputValidationError(
            "demographics_header_duplicate",
            f"{source} 表头不能包含重复列名。",
        )
    return header


def _validated_row(
    header: list[str], values: list[Any], *, row_number: int, source: str
) -> dict[str, Any]:
    if len(values) != len(header):
        raise InputValidationError(
            "demographics_row_width_invalid",
            f"{source} 数据行列数与表头不一致。",
            row_number=row_number,
            expected_columns=len(header),
            actual_columns=len(values),
        )
    missing_columns = [
        column
        for column, value in zip(header, values, strict=True)
        if value is None or (isinstance(value, str) and not value.strip())
    ]
    if missing_columns:
        raise InputValidationError(
            "demographics_cell_missing",
            f"{source} 数据行包含缺失单元格。",
            row_number=row_number,
            columns=missing_columns,
        )
    return dict(zip(header, values, strict=True))


def _read_delimited(path: Path, encoding: str) -> list[dict[str, Any]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        with path.open("r", encoding=encoding, newline="") as stream:
            reader = csv.reader(stream, delimiter=delimiter, strict=True)
            try:
                header = _validated_header(next(reader), source=path.suffix.upper()[1:])
            except StopIteration as exc:
                raise InputValidationError(
                    "demographics_header_missing", "人口学文件缺少表头。"
                ) from exc
            return [
                _validated_row(
                    header,
                    row,
                    row_number=row_number,
                    source=path.suffix.upper()[1:],
                )
                for row_number, row in enumerate(reader, start=2)
            ]
    except UnicodeDecodeError as exc:
        raise InputValidationError(
            "demographics_encoding_error",
            "人口学文件无法使用声明的编码读取。",
            encoding=encoding,
        ) from exc
    except csv.Error as exc:
        raise InputValidationError(
            "demographics_delimited_invalid",
            "CSV/TSV 文件格式无效。",
            reason=str(exc),
        ) from exc


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - dependency gate
        raise InputValidationError(
            "xlsx_support_unavailable", "读取 XLSX 需要安装 openpyxl。"
        ) from exc
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=False)

        def explicit_values(row: tuple[Any, ...]) -> list[Any]:
            """Discard XLSX padding while preserving explicitly blank cells."""

            cells = list(row)
            while cells and getattr(cells[-1], "coordinate", None) is None:
                cells.pop()
            return [cell.value for cell in cells]

        try:
            header = _validated_header(explicit_values(next(rows)), source="XLSX")
        except StopIteration as exc:
            raise InputValidationError("demographics_empty", "人口学文件为空。") from exc
        return [
            _validated_row(
                header,
                explicit_values(row),
                row_number=row_number,
                source="XLSX",
            )
            for row_number, row in enumerate(rows, start=2)
        ]
    finally:
        if "workbook" in locals():
            workbook.close()


def read_demographics(
    path: Path,
    *,
    subject_id_column: str,
    column_mapping: dict[str, str],
    encoding: str,
    manifest_subject_ids: set[str],
) -> dict[str, Any]:
    subject_id_column = subject_id_column.strip()
    if not subject_id_column:
        raise InputValidationError("subject_id_column_invalid", "受试者 ID 列名不能为空。")
    normalized_mapping: dict[str, str] = {}
    used_sources: set[str] = set()
    for target, source in column_mapping.items():
        normalized_target = target.strip()
        normalized_source = source.strip()
        if not normalized_target or not normalized_source:
            raise InputValidationError(
                "demographics_mapping_invalid", "人口学字段映射名称不能为空。"
            )
        if normalized_target == "subject_id":
            raise InputValidationError(
                "demographics_subject_id_reserved",
                "subject_id 是受保护的规范标识, 不能由字段映射覆盖。",
            )
        if normalized_target in normalized_mapping or normalized_source in used_sources:
            raise InputValidationError(
                "demographics_mapping_ambiguous",
                "人口学字段映射在去除首尾空白后必须具有唯一目标和唯一来源。",
            )
        normalized_mapping[normalized_target] = normalized_source
        used_sources.add(normalized_source)
    column_mapping = normalized_mapping
    if path.suffix.lower() in {".csv", ".tsv"}:
        source_rows = _read_delimited(path, encoding)
    elif path.suffix.lower() == ".xlsx":
        source_rows = _read_xlsx(path)
    else:
        raise InputValidationError(
            "demographics_format_unsupported", "人口学文件只支持 CSV、TSV 或 XLSX。"
        )
    if not source_rows:
        raise InputValidationError("demographics_empty", "人口学文件没有数据行。")
    if subject_id_column not in source_rows[0]:
        raise InputValidationError(
            "subject_id_column_missing",
            "人口学文件缺少指定的受试者 ID 列。",
            subject_id_column=subject_id_column,
        )
    selected_sources = {subject_id_column, *column_mapping.values()}
    missing_columns = selected_sources.difference(source_rows[0])
    if missing_columns:
        raise InputValidationError(
            "demographics_columns_missing",
            "人口学字段映射引用了不存在的列。",
            missing_columns=sorted(missing_columns),
        )

    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for source_row in source_rows:
        subject_id = str(source_row.get(subject_id_column, "")).strip()
        if not subject_id:
            raise InputValidationError("subject_id_empty", "人口学文件包含空受试者 ID。")
        if subject_id in seen:
            raise InputValidationError(
                "subject_id_duplicate",
                "人口学文件包含重复受试者 ID。",
                subject_id=subject_id,
            )
        seen.add(subject_id)
        mapped: dict[str, Any] = {"subject_id": subject_id}
        for target, source in sorted(column_mapping.items()):
            value = source_row.get(source)
            mapped[target] = value.strip() if isinstance(value, str) else value
        rows.append(mapped)

    rows.sort(key=lambda row: row["subject_id"])
    return {
        "source_sha256": sha256_file(path),
        "rows": rows,
        "columns": ["subject_id", *sorted(column_mapping)],
        "missing_subject_ids": sorted(manifest_subject_ids.difference(seen)),
        "extra_subject_ids": sorted(seen.difference(manifest_subject_ids)),
    }
