"""CSV/TSV/XLSX inspection without mutating the uploaded file."""

from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]

from neuroagent.analysis.models import TableColumnProfile, TableInspection


def _content_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_rows(path: Path, max_rows: int) -> tuple[list[str], list[list[str]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except (BadZipFile, InvalidFileException) as exc:
            raise ValueError(f"corrupted or invalid XLSX file: {exc}") from exc
        sheet = workbook.active
        values = []
        formula_cells: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=False)):
            if row_index >= max_rows + 1:
                break
            cells: list[str] = []
            for _column_index, cell in enumerate(row, start=1):
                if cell.data_type == "f":
                    formula_cells.append(f"{cell.coordinate}")
                cells.append(_normalize_cell(cell.value))
            values.append(cells)
        workbook.close()
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        raw = path.read_text(encoding="utf-8-sig", errors="strict")
        values = [list(row) for row in csv.reader(raw.splitlines(), delimiter=delimiter)][
            : max_rows + 1
        ]
        formula_cells = []
    else:
        raise ValueError("only CSV, TSV and XLSX tables are supported")
    if not values:
        return [], [], formula_cells
    headers = [_normalize_cell(value) for value in values[0]]
    rows = [[_normalize_cell(value) for value in row] for row in values[1:]]
    return headers, rows, formula_cells


def _column_kind(values: list[str]) -> str:
    nonempty = [value for value in values if value]
    if not nonempty:
        return "empty"
    numeric = 0
    for value in nonempty:
        try:
            float(value)
        except ValueError:
            continue
        numeric += 1
    if numeric == len(nonempty):
        return "numeric"
    if len(set(nonempty)) <= max(20, len(nonempty) // 2):
        return "categorical"
    return "text"


def inspect_table(path: Path, *, max_rows: int = 100_000) -> TableInspection:
    path = path.resolve(strict=True)
    headers, rows, formula_cells = _read_rows(path, max_rows)
    warnings: list[str] = []
    blockers: list[str] = []
    if not headers or not any(headers):
        blockers.append("table_header_missing")
    if len(headers) != len(set(headers)):
        blockers.append("table_header_duplicate")
    if any(value.startswith(("=", "+", "-", "@")) for row in rows for value in row if value):
        warnings.append("formula_like_cell_present")
    if formula_cells:
        warnings.append("xlsx_formula_cells_present")
    normalized_rows = [tuple(row + [""] * (len(headers) - len(row))) for row in rows]
    duplicate_rows = len(normalized_rows) - len(set(normalized_rows))
    if duplicate_rows:
        warnings.append("duplicate_rows_present")
    columns: list[TableColumnProfile] = []
    for index, name in enumerate(headers):
        values = [row[index] if index < len(row) else "" for row in rows]
        nonempty = [value for value in values if value]
        columns.append(
            TableColumnProfile(
                name=name or f"column_{index + 1}",
                kind=_column_kind(values),
                missing_count=len(values) - len(nonempty),
                unique_count=len(set(nonempty)),
                examples=tuple(nonempty[:3]),
            )
        )
    lowered = {name.lower(): name for name in headers if name}
    target_candidates = tuple(
        name
        for lower, name in lowered.items()
        if re.search(r"target|label|group|class|outcome|diagnos", lower)
    )
    subject_candidates = tuple(
        name
        for lower, name in lowered.items()
        if re.search(r"subject|participant|sub[_-]?id|session", lower)
    )
    leakage_candidates = tuple(
        name
        for lower, name in lowered.items()
        if re.search(r"post|result|prediction|fold|test|diagnos", lower)
    )
    if not rows:
        blockers.append("table_empty")
    return TableInspection(
        filename=path.name,
        format={".csv": "csv", ".tsv": "tsv", ".xlsx": "xlsx"}[path.suffix.lower()],
        row_count=len(rows),
        columns=tuple(columns),
        duplicate_rows=max(0, duplicate_rows),
        target_candidates=target_candidates,
        subject_candidates=subject_candidates,
        leakage_candidates=leakage_candidates,
        warnings=tuple(dict.fromkeys(warnings)),
        blocking_issues=tuple(dict.fromkeys(blockers)),
        content_hash=_content_hash(path),
    )
